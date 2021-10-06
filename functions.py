from openpyxl.styles import PatternFill, Font
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl import load_workbook, Workbook
import glob

main_name = "Report.xlsx"

#  Check if main report exists
if main_name not in glob.glob("*.xlsx"):
    main_wb = Workbook()
else:
    main_wb = load_workbook(main_name)

def apply_conditional_format(worksheet, range: str, redmin: str, redmax: str, yellowmin: str, yellowmax: str,
                             greenmin: str):
    redFill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    redFont = Font(color="9C0006")
    yellowFill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    yellowFont = Font(color="9C5700")
    greenFill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    greenFont = Font(color="006100")

    #  	>0 and <0.4 red
    worksheet.conditional_formatting.add(range,
                                         CellIsRule(operator='between', formula=[redmin, redmax], stopIfTrue=True,
                                                    fill=redFill, font=redFont))

    #  	>=0.4 and <0.7 yellow
    worksheet.conditional_formatting.add(range,
                                         CellIsRule(operator='between', formula=[yellowmin, yellowmax], stopIfTrue=True,
                                                    fill=yellowFill, font=yellowFont))

    #  	>=0.7 light green
    worksheet.conditional_formatting.add(range,
                                         CellIsRule(operator='greaterThan', formula=[greenmin], stopIfTrue=True,
                                                    fill=greenFill, font=greenFont))


#  returns most recent raw report as a workbook
def get_most_recent_report():
    import glob
    import os

    files = glob.glob("raw_reports/*.xlsx")
    files.sort(key=os.path.getmtime)
    wb = load_workbook(files[-1])

    return wb



def check_new_raw_reports():

    #  checks for [1] new raw report and if there is, adds it as a sheet to the main report
    most_recent_wb = get_most_recent_report()
    most_recent_worksheet = most_recent_wb.worksheets[0]

    #  check if workbook name in previous wb is in main report
    ws_title = most_recent_wb.active.title.split("_")[-2]  # get date only
    if ws_title in main_wb.sheetnames:
        print("sheet already done")
    else:
        if "Sheet" in main_wb.sheetnames:
            del main_wb["Sheet"]

        main_wb.create_sheet(ws_title)

        # calculate total number of rows and columns in source excel file
        mr = most_recent_worksheet.max_row
        mc = most_recent_worksheet.max_column

        # copying the cell values from source excel file to destination excel file
        for i in range(1, mr + 1):
            for j in range(1, mc + 1):
                # reading cell value from source excel file
                c = most_recent_worksheet.cell(row=i, column=j)

                # writing the read value to destination excel file
                main_wb[ws_title].cell(row=i, column=j).value = c.value

        # calculate_precision_recall_difference()
        format_sheet(ws_title)

        main_wb.save(main_name)


def format_sheet(worksheet):
    # wb = load_workbook(main_name)
    ws1 = main_wb[worksheet]
    print("Formatting sheet " + ws1.title)

    #  delete col D-M
    ws1.delete_cols(4, 10)

    #  delete col G-I
    ws1.delete_cols(7, 9)

    #  enlarge col A
    for cell in ws1["A"]:
        length = max(len(cell.value) for cell in ws1["A"])
        ws1.column_dimensions["A"].width = length

    #  enlarge col Precions, F Measure
    ws1.column_dimensions["D"].width = 12
    ws1.column_dimensions["F"].width = 12

    #  format cell col D-J - Number
    for row in ws1.iter_rows(min_row=2, min_col=4, max_col=10):
        for cell in row:
            cell.number_format = '0.00'

    #  col A-F format as table
    print("Creating table" + ws1.title + "_table")
    tab = Table(displayName=ws1.title + "_table", ref="A1:" + "H" + str(ws1.max_row))
    style = TableStyleInfo(name="TableStyleLight13", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=False, showColumnStripes=False)
    tab.tableStyleInfo = style
    ws1.add_table(tab)

    #  col D-F (minus headers) conditional formatting
    apply_conditional_format(ws1, "D2:F" + str(ws1.max_row), "0.1", "0.39", "0.4", "0.69", "0.7")

    #  add precision and recall average in col J-K
    ws1.column_dimensions["I"].width = 16

    ws1["I2"] = "Precision Average"
    ws1["J2"] = f"=AVERAGE(D2:D{str(ws1.max_row)})"

    ws1["I3"] = "Recall Average"
    ws1["J3"] = f"=AVERAGE(E2:E{str(ws1.max_row)})"



#  most recent precision and recall - (subtract) 2nd most recent precision and recall
def calculate_precision_recall_difference():
    #  get 2nd to last wb if there are more than 1
    print("Calculating precision and recall differences")
    if len(main_wb.sheetnames) > 1:
        last_sheet = main_wb[main_wb.sheetnames[-2]]
        print("older report sheet: " + last_sheet.title)
        now_sheet = main_wb[main_wb.sheetnames[-1]]
        print("newest report sheet: " + now_sheet.title)



        #  paste formula in now p&r difference
        for row in range(2, now_sheet.max_row +1):
            for col in range(7, 8):
                now_precision = now_sheet[f'D{row}'].value
                last_precision = last_sheet[f'D{row}'].value
                if now_precision is None:
                    now_precision = 0
                if last_precision is None:
                    last_precision = 0
                # print(f"np {now_precision}")
                # print(f"lp {last_precision}")
                _ = now_sheet.cell(column=col, row=row, value=f"=SUM({now_precision}, -{last_precision})")
            for col in range(8, 9):
                now_recall = now_sheet[f'E{row}'].value
                last_recall = last_sheet[f'E{row}'].value
                if now_recall is None:
                    now_recall = 0
                if last_recall is None:
                    last_recall = 0
                _ = now_sheet.cell(column=col, row=row, value=f"=SUM({now_recall}, -{last_recall})")

        #  Differential with previous report
        now_sheet.column_dimensions["G"].width = 16
        now_sheet["G1"] = "Precision Diff"
        now_sheet.column_dimensions["H"].width = 16
        now_sheet["H1"] = "Recall Diff"
        apply_conditional_format(now_sheet, "G2:H" + str(now_sheet.max_row), "-999", "-0.01", "0", "0", "0.01")

        main_wb.save(main_name)

    else:
        print("There is only 1 sheet")
